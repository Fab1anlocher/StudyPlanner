""" 
Export Service
Bietet Export-Formate für den Lernplan (PDF, Excel)
"""

from datetime import datetime, date as date_type, timedelta
from typing import List, Dict, Any, Optional
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from constants import WEEKDAY_EN_CAPITALIZED_TO_DE, DATE_FORMAT_DISPLAY


def _sanitize_text_for_pdf(text: str) -> str:
    """
    Sanitize text for PDF output by replacing problematic Unicode characters
    with ASCII equivalents that Helvetica can handle.
    
    Args:
        text: Input text that may contain special characters
        
    Returns:
        Sanitized text safe for PDF rendering
    """
    if not text:
        return text
    
    # Replace common problematic Unicode characters with ASCII equivalents
    replacements = {
        '–': '-',  # en-dash
        '—': '-',  # em-dash
        ''': "'",  # right single quote
        ''': "'",  # left single quote
        '"': '"',  # left double quote
        '"': '"',  # right double quote
        '…': '...',  # ellipsis
        '•': '*',  # bullet
        '→': '->',  # arrow
        '←': '<-',  # arrow
        '↔': '<->',  # double arrow
        '≥': '>=',  # greater or equal
        '≤': '<=',  # less or equal
        '≠': '!=',  # not equal
        '×': 'x',  # multiplication
        '÷': '/',  # division
        '±': '+/-',  # plus-minus
        '€': 'EUR',  # euro
        '£': 'GBP',  # pound
        '¥': 'JPY',  # yen
        '©': '(c)',  # copyright
        '®': '(R)',  # registered
        '™': '(TM)',  # trademark
        '\u00a0': ' ',  # non-breaking space
    }
    
    result = text
    for char, replacement in replacements.items():
        result = result.replace(char, replacement)
    
    return result


def create_pdf_export(
    plan: List[Dict[str, Any]], 
    study_start: date_type, 
    study_end: Optional[date_type],
    busy_times: Optional[List[Dict[str, Any]]] = None,
    absences: Optional[List[Dict[str, Any]]] = None,
    preferences: Optional[Dict[str, Any]] = None,
) -> bytes:
    """
    Create a PDF document from the study plan.

    Args:
        plan: List of study session dicts, sorted chronologically
        study_start: Start date of study plan
        study_end: End date of study plan
        busy_times: Optional list of recurring busy time blocks
        absences: Optional list of absence periods
        preferences: Optional dict with user preferences (limits, etc.)

    Returns:
        PDF file as bytes
    """
    busy_times = busy_times or []
    absences = absences or []
    preferences = preferences or {}
    
    # Create PDF object
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Title
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 10, _sanitize_text_for_pdf("Personalisierter Lernplan"), ln=True, align="C")
    pdf.ln(5)

    # Study plan dates
    pdf.set_font("Helvetica", "", 12)
    start_str = study_start.strftime(DATE_FORMAT_DISPLAY)
    end_str = study_end.strftime(DATE_FORMAT_DISPLAY) if study_end else "N/A"
    pdf.cell(0, 8, _sanitize_text_for_pdf(f"Lernplan: {start_str} - {end_str}"), ln=True, align="C")
    pdf.ln(5)
    
    # Limits info section
    if preferences:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 6, "Lern-Limits:", ln=True)
        pdf.set_font("Helvetica", "", 10)
        
        max_hours_day = preferences.get("max_hours_day", "Kein Limit")
        max_hours_week = preferences.get("max_hours_week", "Kein Limit")
        min_session = preferences.get("min_session_duration", 60)
        
        pdf.cell(0, 5, _sanitize_text_for_pdf(f"  - Max. Stunden/Tag: {max_hours_day}"), ln=True)
        pdf.cell(0, 5, _sanitize_text_for_pdf(f"  - Max. Stunden/Woche: {max_hours_week if max_hours_week else 'Kein Limit'}"), ln=True)
        pdf.cell(0, 5, _sanitize_text_for_pdf(f"  - Min. Session-Dauer: {min_session} Minuten"), ln=True)
        pdf.ln(3)

    # Absences section
    if absences:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 6, "Abwesenheiten:", ln=True)
        pdf.set_font("Helvetica", "", 10)
        
        for absence in absences:
            start_date = absence.get("start_date")
            end_date = absence.get("end_date")
            label = absence.get("label", "Abwesenheit")
            
            if start_date and end_date:
                start_str = start_date.strftime("%d.%m.%Y")
                end_str = end_date.strftime("%d.%m.%Y")
                pdf.cell(0, 5, _sanitize_text_for_pdf(f"  - {label}: {start_str} - {end_str}"), ln=True)
        pdf.ln(3)
    
    # Busy times section
    if busy_times:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 6, "Wiederkehrende Verpflichtungen:", ln=True)
        pdf.set_font("Helvetica", "", 10)
        
        for busy in busy_times:
            label = busy.get("label", "Belegt")
            days = ", ".join(busy.get("days", []))
            start = busy.get("start", "")
            end = busy.get("end", "")
            pdf.cell(0, 5, _sanitize_text_for_pdf(f"  - {label}: {days}, {start}-{end}"), ln=True)
        pdf.ln(3)

    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "Lernplan Details:", ln=True)
    pdf.ln(3)

    # Group sessions by date
    sessions_by_date = {}
    for session in plan:
        date_key = session.get("date", "Unknown")
        if date_key not in sessions_by_date:
            sessions_by_date[date_key] = []
        sessions_by_date[date_key].append(session)

    # Pre-compute busy times by weekday for O(1) lookup
    weekday_names_de = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
    busy_by_weekday = {i: [] for i in range(7)}
    for busy in busy_times:
        for day_name in busy.get("days", []):
            try:
                day_idx = weekday_names_de.index(day_name)
                busy_by_weekday[day_idx].append(busy)
            except ValueError:
                pass

    # Iterate through each date
    for date_str in sorted(sessions_by_date.keys()):
        # Parse and format date with German weekday
        try:
            session_date = datetime.fromisoformat(date_str).date()
            weekday_en = session_date.strftime("%A")
            weekday_de = WEEKDAY_EN_CAPITALIZED_TO_DE.get(weekday_en, weekday_en)
            date_display = f"{weekday_de}, {session_date.strftime(DATE_FORMAT_DISPLAY)}"
            weekday_idx = session_date.weekday()
        except:
            date_display = date_str
            weekday_idx = -1

        # Date header (bold)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, _sanitize_text_for_pdf(date_display), ln=True)
        
        # Show busy times for this weekday
        if weekday_idx >= 0 and busy_by_weekday[weekday_idx]:
            for busy in busy_by_weekday[weekday_idx]:
                label = busy.get("label", "Belegt")
                start = busy.get("start", "")
                end = busy.get("end", "")
                pdf.set_font("Helvetica", "", 9)
                pdf.set_text_color(150, 100, 0)  # Orange-brown color
                pdf.cell(0, 5, _sanitize_text_for_pdf(f"  [Belegt] {start}-{end}: {label}"), ln=True)
                pdf.set_text_color(0, 0, 0)  # Reset to black

        # Sessions for this date
        sessions = sessions_by_date[date_str]

        for session in sessions:
            start = session.get("start", "N/A")
            end = session.get("end", "N/A")
            module = session.get("module", "Unknown")
            topic = session.get("topic", "N/A")
            description = session.get("description", "")

            # Time (bold)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(35, 6, _sanitize_text_for_pdf(f"{start} - {end}"), ln=False)

            # Module
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, _sanitize_text_for_pdf(module), ln=True)

            # Topic (indented)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_x(45)
            pdf.cell(0, 5, _sanitize_text_for_pdf(f"Thema: {topic}"), ln=True)

            # Description (indented, smaller font)
            if description:
                pdf.set_font("Helvetica", "", 8)
                pdf.set_x(45)
                pdf.multi_cell(0, 5, _sanitize_text_for_pdf(description))

            pdf.ln(1)

        pdf.ln(3)

    # Footer with generation info
    pdf.set_y(-30)
    pdf.set_font("Helvetica", "I", 9)
    pdf.cell(
        0,
        5,
        _sanitize_text_for_pdf(f"Erstellt am {datetime.now().strftime('%d.%m.%Y um %H:%M')}"),
        ln=True,
        align="C",
    )
    pdf.cell(0, 5, "KI-Lernplaner", ln=True, align="C")

    # Return PDF as bytes
    return bytes(pdf.output())


def create_excel_export(
    plan: List[Dict[str, Any]],
    leistungsnachweise: List[Dict[str, Any]],
    preferences: Dict[str, Any],
    study_start: date_type,
    study_end: Optional[date_type],
    busy_times: List[Dict[str, Any]],
    absences: List[Dict[str, Any]],
) -> bytes:
    """
    Create an Excel workbook (.xlsx) with study plan and configuration.
    Contains multiple sheets for different data.

    Args:
        plan: Generated study plan
        leistungsnachweise: List of assessments
        preferences: User preferences
        study_start: Start date
        study_end: End date
        busy_times: Recurring busy times
        absences: Absence periods

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def set_date_cell(cell, date_value, apply_border=True):
        """Helper to format a cell as a date with DD.MM.YYYY format."""
        cell.value = date_value
        if date_value:
            cell.number_format = 'DD.MM.YYYY'
        if apply_border:
            cell.border = border
    
    # Sheet 1: Lernplan (Study Sessions)
    ws_plan = wb.active
    ws_plan.title = "Lernplan"
    
    # Headers
    headers = ["Datum", "Wochentag", "Start", "Ende", "Dauer (h)", "Modul", "Thema", "Beschreibung", "Lernmethode"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_plan.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Data rows
    for row_num, session in enumerate(plan, 2):
        date_str = session.get("date", "")
        start_time = session.get("start", "")
        end_time = session.get("end", "")
        
        # Calculate duration
        duration = 0.0
        try:
            start_dt = datetime.strptime(start_time, "%H:%M")
            end_dt = datetime.strptime(end_time, "%H:%M")
            duration = (end_dt - start_dt).total_seconds() / 3600
        except:
            pass
        
        # Parse date and get weekday
        weekday = ""
        session_date = None
        try:
            session_date = datetime.fromisoformat(date_str).date()
            weekday_en = session_date.strftime("%A")
            weekday = WEEKDAY_EN_CAPITALIZED_TO_DE.get(weekday_en, weekday_en)
        except:
            pass
        
        # Write date cell with proper Excel date format using helper
        date_cell = ws_plan.cell(row=row_num, column=1)
        if session_date:
            set_date_cell(date_cell, session_date)
        else:
            date_cell.value = date_str
            date_cell.border = border
        
        # Write remaining data (start from column 2 since date is in column 1)
        row_data = [
            weekday,
            start_time,
            end_time,
            round(duration, 2),
            session.get("module", ""),
            session.get("topic", ""),
            session.get("description", ""),
            session.get("learning_method", "")
        ]
        
        for col_num, value in enumerate(row_data, 2):
            cell = ws_plan.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num in [3, 4]:  # Time columns
                cell.alignment = Alignment(horizontal="center")
            elif col_num == 5:  # Duration column
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '0.00'
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws_plan.column_dimensions[get_column_letter(col)].width = 15
    ws_plan.column_dimensions['H'].width = 40  # Description column wider
    
    # Sheet 2: Leistungsnachweise (Assessments)
    ws_ln = wb.create_sheet("Leistungsnachweise")
    
    ln_headers = ["Titel", "Typ", "Deadline", "Modul", "Themen", "Priorität", "Aufwand", "Prüfungsformat", "Details"]
    for col_num, header in enumerate(ln_headers, 1):
        cell = ws_ln.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    for row_num, ln in enumerate(leistungsnachweise, 2):
        # Serialize exam_format if it's an enum
        exam_format = ln.get("exam_format")
        if exam_format and hasattr(exam_format, "value"):
            exam_format = exam_format.value
        
        # Get type as string
        ln_type = ln.get("type", "")
        if hasattr(ln_type, "value"):
            ln_type = ln_type.value
        
        # Get deadline as date object
        deadline = ln.get("deadline")
        
        topics_str = ", ".join(ln.get("topics", []))
        
        # Write Title
        cell = ws_ln.cell(row=row_num, column=1, value=ln.get("title", ""))
        cell.border = border
        
        # Write Type
        cell = ws_ln.cell(row=row_num, column=2, value=ln_type)
        cell.border = border
        
        # Write Deadline with proper date format using helper
        set_date_cell(ws_ln.cell(row=row_num, column=3), deadline)
        
        # Write remaining data
        remaining_data = [
            ln.get("module", ""),
            topics_str,
            ln.get("priority", 3),
            ln.get("effort", 3),
            exam_format or "",
            ln.get("exam_details", "")
        ]
        
        for col_num, value in enumerate(remaining_data, 4):
            cell = ws_ln.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num in [6, 7]:  # Priority and Effort
                cell.alignment = Alignment(horizontal="center")
    
    for col in range(1, len(ln_headers) + 1):
        ws_ln.column_dimensions[get_column_letter(col)].width = 15
    ws_ln.column_dimensions['A'].width = 25  # Title wider
    ws_ln.column_dimensions['E'].width = 30  # Topics wider
    
    # Sheet 3: Abwesenheiten (Absences)
    ws_abs = wb.create_sheet("Abwesenheiten")
    
    abs_headers = ["Start", "Ende", "Grund"]
    for col_num, header in enumerate(abs_headers, 1):
        cell = ws_abs.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    for row_num, absence in enumerate(absences, 2):
        start_date = absence.get("start_date")
        end_date = absence.get("end_date")
        
        # Write start and end dates with proper format using helper
        set_date_cell(ws_abs.cell(row=row_num, column=1), start_date)
        set_date_cell(ws_abs.cell(row=row_num, column=2), end_date)
        
        # Write label
        cell = ws_abs.cell(row=row_num, column=3, value=absence.get("label", ""))
        cell.border = border
    
    for col in range(1, len(abs_headers) + 1):
        ws_abs.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 4: Verpflichtungen (Busy Times)
    ws_busy = wb.create_sheet("Verpflichtungen")
    
    busy_headers = ["Bezeichnung", "Tage", "Start", "Ende"]
    for col_num, header in enumerate(busy_headers, 1):
        cell = ws_busy.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    for row_num, busy in enumerate(busy_times, 2):
        days_str = ", ".join(busy.get("days", []))
        
        row_data = [
            busy.get("label", ""),
            days_str,
            busy.get("start", ""),
            busy.get("end", "")
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws_busy.cell(row=row_num, column=col_num, value=value)
            cell.border = border
    
    for col in range(1, len(busy_headers) + 1):
        ws_busy.column_dimensions[get_column_letter(col)].width = 20
    
    # Sheet 5: Übersicht (Summary)
    ws_summary = wb.create_sheet("Übersicht")
    ws_summary.sheet_properties.tabColor = "FFD700"  # Golden tab
    
    # Write summary data with proper formatting
    summary_rows = [
        ("Lernplan-Übersicht", "", "header"),
        ("", "", "empty"),
        ("Zeitraum", "", "header"),
        ("Start", study_start, "date"),
        ("Ende", study_end, "date"),
        ("", "", "empty"),
        ("Statistiken", "", "header"),
        ("Gesamt Lernsessions", len(plan), "number"),
        ("Gesamt Lerntage", len(set(s.get("date", "") for s in plan)), "number"),
        ("Anzahl Module", len(set(s.get("module", "") for s in plan)), "number"),
        ("Anzahl Leistungsnachweise", len(leistungsnachweise), "number"),
        ("Anzahl Abwesenheiten", len(absences), "number"),
        ("Anzahl Verpflichtungen", len(busy_times), "number"),
        ("", "", "empty"),
        ("Exportiert am", datetime.now().strftime("%d.%m.%Y %H:%M:%S"), "text"),
    ]
    
    for row_num, (label, value, value_type) in enumerate(summary_rows, 1):
        cell_a = ws_summary.cell(row=row_num, column=1, value=label)
        cell_b = ws_summary.cell(row=row_num, column=2, value=value)
        
        # Format date cells
        if value_type == "date" and value:
            cell_b.number_format = 'DD.MM.YYYY'
        
        if label in ["Lernplan-Übersicht", "Zeitraum", "Statistiken"]:
            cell_a.font = Font(bold=True, size=14)
            cell_a.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        else:
            cell_a.font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def get_plan_statistics(plan: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Calculate statistics about the study plan.

    Args:
        plan: Study plan sessions

    Returns:
        Dictionary with statistics
    """
    if not plan:
        return {
            "total_sessions": 0,
            "total_hours": 0.0,
            "unique_modules": 0,
            "unique_dates": 0,
        }

    # Single pass through the plan to collect all statistics
    total_hours = 0.0
    modules = set()
    dates = set()
    
    for session in plan:
        # Track unique modules and dates
        modules.add(session.get("module", ""))
        dates.add(session.get("date", ""))
        
        # Calculate session duration
        try:
            start = datetime.strptime(session.get("start", "00:00"), "%H:%M")
            end = datetime.strptime(session.get("end", "00:00"), "%H:%M")
            hours = (end - start).total_seconds() / 3600
            total_hours += hours
        except:
            pass

    return {
        "total_sessions": len(plan),
        "total_hours": round(total_hours, 1),
        "unique_modules": len(modules),
        "unique_dates": len(dates),
    }
